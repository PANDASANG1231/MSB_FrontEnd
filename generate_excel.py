
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
import os

def generate_excel_report(csv_filename="fake_data.csv", images_dir="images", output_filename="report.xlsx"):
    # Read the fake data
    df = pd.read_csv(csv_filename)

    # Convert DIN to string to ensure it's treated as text in Excel and for consistent linking
    df["DIN"] = df["DIN"].astype(str)

    # Convert boolean to integer for summation
    df["pred_3rd_party_check"] = df["pred_3rd_party_check"].astype(int)

    # Create summary table
    summary_df = df.groupby("acct_no").agg(
        total_3rd_party_deposit=("pred_3rd_party_check", "sum"),
        row_count=("acct_no", "size")
    ).reset_index()

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Sheet 1: Summary Table
    sheet1 = workbook.active
    sheet1.title = "Summary Table"
    sheet1.sheet_view.showGridLines = False # Remove gridlines

    # Write summary table title
    sheet1.cell(row=1, column=2, value="Summary Table by Account Number")
    
    # Sort summary_df by total_3rd_party_deposit
    summary_df = summary_df.sort_values(by="total_3rd_party_deposit", ascending=False)

    # Sort df by acct_no (This needs to be done before creating the mapping)
    df = df.sort_values(by="acct_no")

    # Create a mapping of acct_no to its first row in the raw data sheet
    acct_no_to_raw_data_row = {}
    for r_idx, row_data in enumerate(df.iterrows(), 3): # Data starts at row 3 in Excel
        acct_no = row_data[1]["acct_no"]
        # Update to store the *last* occurrence
        acct_no_to_raw_data_row[acct_no] = r_idx

    # Write header for summary_df
    for col_idx, value in enumerate(summary_df.columns, 2): # Start from column 2
        cell = sheet1.cell(row=2, column=col_idx, value=value)
        cell.font = Font(bold=True)

    # Write data for summary_df with hyperlinks
    for r_idx, row_data in enumerate(summary_df.iterrows(), 3): # Start from row 3
        for c_idx, value in enumerate(row_data[1], 2): # Start from column 2
            cell = sheet1.cell(row=r_idx, column=c_idx, value=value)
            if summary_df.columns[c_idx - 2] == "acct_no": # Check if it's the acct_no column
                if value in acct_no_to_raw_data_row:
                    # Create hyperlink to the first occurrence of this acct_no in the Raw Data sheet
                    target_row = acct_no_to_raw_data_row[value] # Link to first row + 30
                    cell.hyperlink = f"#'Raw Data'!A{target_row}" # Link to column A (start of the row)
                    cell.font = Font(color="0000FF", underline="single") # Make it look like a link
            

    # Apply borders and adjust column widths for Summary Table
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Define the range of the summary table for borders (adjust as needed)
    # Assuming data starts at B2 and goes up to max_row, max_column of summary_df
    summary_start_row = 2
    summary_end_row = sheet1.max_row
    summary_start_col = 2
    summary_end_col = summary_start_col + len(summary_df.columns) - 1

    for r in range(summary_start_row, summary_end_row + 1):
        for c in range(summary_start_col, summary_end_col + 1):
            sheet1.cell(row=r, column=c).border = thin_border

    for column_cells in sheet1.columns:
        length = max(len(str(cell.value)) for cell in column_cells) #+1 to count header.max length
        sheet1.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    # Create a mapping of acct_no to its row in the summary table
    acct_no_to_summary_row = {}
    for r_idx, row_data in enumerate(summary_df.iterrows(), 3): # Data starts at row 3 in Excel
        acct_no = row_data[1]["acct_no"]
        acct_no_to_summary_row[acct_no] = r_idx

    # Sheet 2: Raw Data
    sheet2 = workbook.create_sheet(title="Raw Data")
    sheet2.sheet_view.showGridLines = False # Remove gridlines

    # Write raw data title
    sheet2.cell(row=1, column=2, value="Raw Fake Data") # Title starts at column 2

    # Sort df by acct_no
    df = df.sort_values(by="acct_no")

    # Create a mapping of acct_no to its first row in the raw data sheet
    acct_no_to_raw_data_row = {}
    for r_idx, row_data in enumerate(df.iterrows(), 3): # Data starts at row 3 in Excel
        acct_no = row_data[1]["acct_no"]
        # Update to store the *last* occurrence
        acct_no_to_raw_data_row[acct_no] = r_idx

    # Create a mapping for image links before it's used
    image_link_mapping = {}
    current_row_for_image = 2 # Start image placement from row 2 on the Images sheet
    for filename in os.listdir(images_dir):
        if filename.endswith(".jpg") or filename.endswith(".png"):
            filepath = os.path.join(images_dir, filename)
            parts = filename.split('-')
            acct_no = parts[0]
            din = parts[1].split('.')[0]
            image_link_mapping[(acct_no, din)] = current_row_for_image
            current_row_for_image += 1 # Increment for the next image's row in the images sheet


    # Write raw data header
    for col_idx, value in enumerate(df.columns, 2): # Start from column 2
        cell = sheet2.cell(row=2, column=col_idx, value=value)
        cell.font = Font(bold=True)

    # Add "Link to Summary" header
    link_col_idx = len(df.columns) + 2 # Column after the last data column
    cell = sheet2.cell(row=2, column=link_col_idx, value="Link to Summary")
    cell.font = Font(bold=True)

    # Create a mapping of (acct_no, DIN) to its first row in the raw data sheet
    acct_no_din_to_raw_data_row = {}
    for r_idx, row_data in enumerate(df.iterrows(), 3):
        acct_no = row_data[1]["acct_no"]
        din = str(row_data[1]["DIN"]) # Ensure DIN is a string to match image_link_mapping key
        if (acct_no, din) not in acct_no_din_to_raw_data_row:
            acct_no_din_to_raw_data_row[(acct_no, din)] = r_idx

    # Write raw data
    for r_idx, row_data in enumerate(df.iterrows(), 3):
        for c_idx, value in enumerate(row_data[1], 2): # Start from column 2
            cell = sheet2.cell(row=r_idx, column=c_idx, value=value)
            
            # If this is the DIN column, add a hyperlink to the image sheet
            if df.columns[c_idx - 2] == "DIN":
                acct_no = row_data[1]["acct_no"]
                din = str(value) # DIN is an integer, convert to string for key
                if (acct_no, din) in image_link_mapping:
                    target_image_row = image_link_mapping[(acct_no, din)]
                    cell.hyperlink = f"#'Images'!A{target_image_row}" # Link to column A of the image row
                    cell.font = Font(color="0000FF", underline="single")
        
        # Add hyperlink to summary table
        acct_no = row_data[1]["acct_no"]
        if acct_no in acct_no_to_summary_row:
            target_row_summary = acct_no_to_summary_row[acct_no]
            cell = sheet2.cell(row=r_idx, column=link_col_idx, value="Link to Summary")
            cell.hyperlink = f"#'Summary Table'!B{target_row_summary}" # Link to column B (acct_no column)
            cell.font = Font(color="0000FF", underline="single")

    # Apply borders and adjust column widths for Raw Data Table
    raw_data_start_row = 2
    raw_data_end_row = sheet2.max_row
    raw_data_start_col = 2 # Table starts at column 2
    raw_data_end_col = raw_data_start_col + len(df.columns) - 1

    for r in range(raw_data_start_row, raw_data_end_row + 1):
        for c in range(raw_data_start_col, raw_data_end_col + 1):
            sheet2.cell(row=r, column=c).border = thin_border

    # Set uniform wide width for columns A, B, C, D, E in Raw Data sheet
    for col_idx in range(1, 6): # Columns A to E
        sheet2.column_dimensions[get_column_letter(col_idx)].width = 25 # Set a wide width, adjust as needed

    # Removed auto-sizing for raw data columns to apply uniform width
    # for column_cells in sheet2.columns:
    #     length = max(len(str(cell.value)) for cell in column_cells) #+1 to count header.max length
    #     sheet2.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    # Sheet 3: Images
    sheet3 = workbook.create_sheet(title="Images")
    sheet3.sheet_view.showGridLines = False # Remove gridlines

    # Set column headers for Images sheet
    sheet3.cell(row=1, column=1, value="Account No").font = Font(bold=True)
    sheet3.cell(row=1, column=2, value="DIN").font = Font(bold=True)
    sheet3.cell(row=1, column=3, value="Link to Detail").font = Font(bold=True) # Now in Column C
    sheet3.cell(row=1, column=4, value="Image").font = Font(bold=True) # Now in Column D

    # Set initial column widths for A, B, C, D
    sheet3.column_dimensions[get_column_letter(1)].width = 20 # Column A for Account No
    sheet3.column_dimensions[get_column_letter(2)].width = 20 # Column B for DIN
    sheet3.column_dimensions[get_column_letter(3)].width = 20 # Column C for Link to Detail
    sheet3.column_dimensions[get_column_letter(4)].width = 50 # Column D for Image (adjust as needed)

    # Initialize current_row_for_image here
    current_row_for_image = 2 # Start image placement from row 2
    image_link_mapping = {}
    
    for filename in os.listdir(images_dir):
        if filename.endswith(".jpg") or filename.endswith(".png"):
            filepath = os.path.join(images_dir, filename)
            
            # Extract acct_no and DIN from filename (e.g., acct_no-DIN.jpg)
            parts = filename.split('-')
            acct_no = parts[0]
            din = parts[1].split('.')[0]

            # Store the row where this image will be placed
            # The mapping still refers to where the image *is*, which is now col D
            image_link_mapping[(acct_no, din)] = current_row_for_image

            img = Image(filepath)
            # Adjust image size if necessary (optional)
            img.width = 400
            img.height = 300

            # Add acct_no and DIN to columns A and B
            sheet3.cell(row=current_row_for_image, column=1, value=acct_no)
            sheet3.cell(row=current_row_for_image, column=2, value=din)

            # Add 'Link to Detail' hyperlink to column C
            if (acct_no, din) in acct_no_din_to_raw_data_row:
                target_raw_data_row = acct_no_din_to_raw_data_row[(acct_no, din)]
                cell = sheet3.cell(row=current_row_for_image, column=3, value="Link to Detail") # Column 3 is 'Link to Detail'
                cell.hyperlink = f"#'Raw Data'!A{target_raw_data_row}" # Link to column A of the detail row
                cell.font = Font(color="0000FF", underline="single")

            # Add image to column D
            img.anchor = get_column_letter(4) + str(current_row_for_image) # Anchor image to column D
            sheet3.add_image(img)

            # Adjust row height to accommodate the image
            sheet3.row_dimensions[current_row_for_image].height = img.height * 0.75 # Approximation to fit image height
            
            current_row_for_image += 1 # Move to the next row for the next image

    # Save the workbook
    workbook.save(output_filename)
    print(f"Excel report generated: {output_filename}")

if __name__ == "__main__":
    generate_excel_report()
